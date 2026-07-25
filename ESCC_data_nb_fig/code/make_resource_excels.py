#!/usr/bin/env python3
import argparse
import csv
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parent.parent

# Short, publication-facing sheet names and plain-English descriptions.
# Keys are core-table stems so the filenames can remain stable and auditable.
TABLE_META = {
    "Figure_1b_WES_scRNA_tumor_sample_membership": (
        "Fig1b_membership",
        "Tumor samples included in WES, scRNA-seq, or both datasets.",
    ),
    "Figure_1b_scRNA_patient_sampling_counts": (
        "Fig1b_sampling",
        "Patient counts by response group and pre/post-treatment sampling pattern.",
    ),
    "Figure_1d_major_cell_type_proportions": (
        "Fig1d",
        "Major cell-type percentages in each tumor sample.",
    ),
    "Figure_1e_cell_type_tree_counts": (
        "Fig1e_counts",
        "Cell counts used to build the cell-type hierarchy.",
    ),
    "Figure_1e_subtype_response_composition": (
        "Fig1e_response",
        "Cell-subtype composition by treatment-response group.",
    ),
    "Figure_1e_subtype_sample_type_composition": (
        "Fig1e_timepoint",
        "Cell-subtype composition before and after treatment.",
    ),
    "Figure_1e_tree_edges": (
        "Fig1e_edges",
        "Parent-child links defining the cell-type hierarchy.",
    ),
    "Figure_2a_b_clonal_mutation_counts": (
        "Fig2a-b",
        "Clonal mutation counts per sample for the Figure 2a and 2b comparisons.",
    ),
    "Figure_2c_expression_weighted_neoantigen_counts": (
        "Fig2c",
        "Expression-weighted neoantigen burden per sample.",
    ),
    "Figure_2d_immune_editing_scores": (
        "Fig2d",
        "Immune-editing scores for the analyzed samples.",
    ),
    "Figure_2e_epithelial_tumor_normal_proportions": (
        "Fig2e",
        "Percentages of tumor-like and normal epithelial cells in each sample.",
    ),
    "Figure_2f_epithelial_cluster_frequency_by_patient": (
        "Fig2f",
        "Epithelial-cluster percentages by patient used in the heatmap.",
    ),
    "Figure_2g_epithelial_marker_dotplot": (
        "Fig2g",
        "Average expression and detection rate of epithelial-cluster markers.",
    ),
    "Figure_2h_Gp1_Gp2_epithelial_proportions": (
        "Fig2h",
        "Tumor Gp1, tumor Gp2, and normal epithelial-cell percentages by sample.",
    ),
    "Figure_2i_Gp1_percentage_vs_clonal_mutation_count": (
        "Fig2i",
        "Tumor Gp1 percentage and clonal mutation count for each matched sample.",
    ),
    "Figure_2j_CellChat_network_edges": (
        "Fig2j_edges",
        "Filtered cell-cell communication links shown in the network.",
    ),
    "Figure_2j_CellChat_network_nodes": (
        "Fig2j_nodes",
        "Cell groups shown as nodes in the communication network.",
    ),
    "Figure_2k_ECGEA_logrank_statistics": (
        "Fig2k_log-rank",
        "Global log-rank test for the three ECGEA marker-score groups.",
    ),
    "Figure_2k_ECGEA_survival_marker_stratification": (
        "Fig2k_survival",
        "Patient survival time, event status, and ECGEA marker-score group.",
    ),
    "Figure_2l_Gp1_change_nonresponder_groups": (
        "Fig2l",
        "Tumor Gp1 percentages and detailed non-responder subgroup assignments.",
    ),
    "Figure_2m_Treg_expanded_cell_proportions": (
        "Fig2m",
        "Expanded-cell percentages in regulatory T-cell subtypes.",
    ),
    "Figure_2n_virus_detected_cell_percentages": (
        "Fig2n",
        "Percentage of virus-detected cells in each sample.",
    ),
    "Figure_2o_virus_detected_Gp1_Gp2_percentages": (
        "Fig2o",
        "Virus-detected cell percentages in tumor Gp1 and Gp2 epithelial cells.",
    ),
    "Figure_3a_T_NK_marker_dotplot": (
        "Fig3a",
        "Average expression and detection rate of T- and NK-cell markers.",
    ),
    "Figure_3b_TCR_sharing_edges": (
        "Fig3b_edges",
        "TCR-sharing links between T-cell clusters.",
    ),
    "Figure_3b_TCR_sharing_nodes": (
        "Fig3b_nodes",
        "T-cell clusters shown in the TCR-sharing network.",
    ),
    "Figure_3c_selected_T_cell_cluster_proportions": (
        "Fig3c",
        "Selected T-cell cluster percentages in each sample.",
    ),
    "Figure_3d_CXCL13_CD8_mIF_by_image_field": (
        "Fig3d_fields",
        "CXCL13-positive CD8-cell fractions for individual multiplex-imaging fields.",
    ),
    "Figure_3d_CXCL13_CD8_mIF_patient_summary": (
        "Fig3d_patients",
        "Patient-level summary of CXCL13-positive CD8-cell fractions.",
    ),
    "Figure_3e_CXCR4_CXCL12_major_cell_dotplot": (
        "Fig3e_major",
        "CXCR4 and CXCL12 expression across major cell types.",
    ),
    "Figure_3e_CXCR4_T_cell_dotplot": (
        "Fig3e_T_cells",
        "CXCR4 expression across T-cell clusters.",
    ),
    "Figure_3f_CXCL12_endothelial_expression": (
        "Fig3f_endothelial",
        "CXCL12 expression in endothelial cells for each sample.",
    ),
    "Figure_3f_CXCL12_fibroblast_expression": (
        "Fig3f_fibroblast",
        "CXCL12 expression in fibroblasts for each sample.",
    ),
    "Figure_3g_CXCR4_cytotoxic_CD8_expression": (
        "Fig3g_cytotoxic",
        "CXCR4 expression in cytotoxic CD8 T cells for each sample.",
    ),
    "Figure_3g_CXCR4_exhausted_CD8_expression": (
        "Fig3g_exhausted",
        "CXCR4 expression in exhausted CD8 T cells for each sample.",
    ),
    "Figure_4a_post_novel_vs_preexisting_clonotype_fraction": (
        "Fig4a",
        "Post-treatment fractions of novel and pre-existing clonotypes.",
    ),
    "Figure_4b_c_shared_clonotypes_in_posttreatment": (
        "Fig4b-c_post",
        "Shared-clonotype percentages in post-treatment samples.",
    ),
    "Figure_4b_shared_clonotypes_in_pretreatment": (
        "Fig4b_pre",
        "Shared-clonotype percentages in pre-treatment samples.",
    ),
    "Figure_4d_pre_post_shared_clonotype_fractions": (
        "Fig4d",
        "Pre- and post-treatment fractions of shared clonotypes by cell state.",
    ),
    "Figure_4e_clone_change_type_proportions": (
        "Fig4e",
        "Expanded- and contracted-clonotype percentages by cell state.",
    ),
    "Figure_4f_clone_category_cell_type_composition": (
        "Fig4f",
        "Cell-type composition of cytotoxic-expanded and exhausted-contracted clones.",
    ),
    "Figure_4g_exhausted_cluster_proportions_in_contracted_clones": (
        "Fig4g",
        "Exhausted-cluster percentages within contracted clonotypes.",
    ),
    "Figure_4h_exhausted_cluster_proportions_in_expanded_clones": (
        "Fig4h",
        "Exhausted-cluster percentages within cytotoxic expanded clonotypes.",
    ),
    "Figure_5a_expansion_change_by_patient": (
        "Fig5a_values",
        "Pre-to-post clonotype-expansion changes for each patient and T-cell subtype.",
    ),
    "Figure_5a_expansion_change_statistics": (
        "Fig5a_stats",
        "Between-group statistics for clonotype-expansion changes.",
    ),
    "Figure_5b_CX3CR1_clonotype_diversity": (
        "Fig5b_diversity",
        "CD8 CX3CR1 clonotype diversity for each sample.",
    ),
    "Figure_5b_CX3CR1_expanded_cell_proportions": (
        "Fig5b_expansion",
        "Expanded-cell percentages in the CD8 CX3CR1 cluster.",
    ),
    "Figure_5c_cytotoxicity_exhaustion_scores": (
        "Fig5c",
        "Single-cell cytotoxicity and exhaustion scores by clonotype-sharing group.",
    ),
    "Figure_6a_in_vitro_killing_efficiency": (
        "Fig6a",
        "In vitro tumor-killing efficiency for the four CD8 T-cell populations.",
    ),
    "Figure_6e_OT1_CX3CR1_flow_cytometry_ratios": (
        "Fig6e",
        "Flow-cytometry ratios for OT1 and CX3CR1-positive CD8 T cells.",
    ),
}


parser = argparse.ArgumentParser()
parser.add_argument("--figure", choices=[f"Figure{i}" for i in range(1, 7)])
args = parser.parse_args()

src = ROOT / "resource_data" / "core_tables" / args.figure
out = ROOT / "resource_data" / f"{args.figure}_resource_data.xlsx"
tables = sorted(src.glob("*.tsv"))

missing_meta = [table.stem for table in tables if table.stem not in TABLE_META]
if missing_meta:
    raise KeyError(f"Missing sheet metadata: {', '.join(missing_meta)}")

titles = [TABLE_META[table.stem][0] for table in tables]
if len(titles) != len(set(titles)):
    raise ValueError(f"Duplicate sheet names for {args.figure}")
if any(len(title) > 31 for title in titles):
    raise ValueError("Excel sheet names must be 31 characters or fewer")

wb = Workbook(write_only=True)
contents = wb.create_sheet("Contents")
contents.append(["sheet", "description"])

for table in tables:
    title, description = TABLE_META[table.stem]
    contents.append([title, description])
    ws = wb.create_sheet(title)
    with table.open(encoding="utf-8", newline="") as fh:
        for row in csv.reader(fh, delimiter="\t"):
            ws.append(row)

out.parent.mkdir(exist_ok=True)
wb.save(out)
print(out)
